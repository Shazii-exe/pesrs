"""
run_eval.py
───────────────────────────────────────────────────────────────
PEISR Evaluation Runner — generates paper-ready results.

Usage:
    python run_eval.py                   # runs all 75 prompts
    python run_eval.py --limit 10        # quick smoke test (first 10)
    python run_eval.py --route TECH      # only TECH prompts
    python run_eval.py --quality messy   # only messy prompts

Outputs:
    peisr_eval_results.xlsx   ← main results sheet (open in Excel)
    peisr_eval_summary.txt    ← summary stats for paper

What it measures:
    1. Gate accuracy       — did the gate correctly pass/rewrite?
    2. Prompt quality lift — score before vs after rewrite
    3. Response win rate   — how often does enhanced beat baseline?
    4. Rewrite rate        — % of prompts that triggered rewrite per route
    5. Latency             — avg time per route
───────────────────────────────────────────────────────────────
"""

from __future__ import annotations

import argparse
import json
import time
import traceback
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from answerer import run_pipeline
from judge import judge_pair, consistent_judge_pair, heuristic_judge_pair, total_score
from eval_prompts import EVAL_PROMPTS
import gemini_client as _gc


# ── Excel output path ──────────────────────────────────────────
XLSX_OUT = "peisr_eval_results.xlsx"
TXT_OUT  = "peisr_eval_summary.txt"

# ── Column headers for results sheet ──────────────────────────
HEADERS = [
    "prompt_id", "route_expected", "route_predicted", "route_correct",
    "quality_tier", "expected_gate", "actual_gate", "gate_correct",
    "original_prompt", "enhanced_prompt", "rewritten",
    "orig_score_total", "orig_intent", "orig_clarity", "orig_structure", "orig_safety",
    "final_score_total", "final_intent", "final_clarity", "final_structure", "final_safety",
    "score_lift",
    "threshold_used", "threshold_reason",
    "enhance_mode", "temperature_used",
    "baseline_response", "enhanced_response",
    "llm_judge_winner", "llm_judge_reason",
    "llm_score_x_total", "llm_score_y_total",
    "heur_judge_winner",
    "judges_agree",
    "judge_consistent",
    "llm_winner_run2",
    "model_used",
    "latency_seconds",
    "error",
]

# ── Colors ─────────────────────────────────────────────────────
COL_HEADER  = "1e1b4b"
COL_PASS    = "d1fae5"
COL_FAIL    = "fee2e2"
COL_WARN    = "fef9c3"
COL_NEUTRAL = "f3f4f6"


def _cell_color(ws, row, col, hex_color):
    ws.cell(row, col).fill = PatternFill("solid", fgColor=hex_color)


def run_single(prompt_data: dict) -> dict:
    """Run one prompt through the full PEISR pipeline and return a result row."""
    pid      = prompt_data["id"]
    text     = prompt_data["text"]
    route_ex = prompt_data["route"]
    quality  = prompt_data["quality"]
    exp_gate = prompt_data["expected_gate"]

    result = {
        "prompt_id": pid,
        "route_expected": route_ex,
        "quality_tier": quality,
        "expected_gate": exp_gate,
        "original_prompt": text,
        "error": "",
    }

    t0 = time.time()

    try:
        # Run ABC pipeline (gated)
        out = run_pipeline(text, variant="ABC", temp_mode="auto", max_rounds=1)

        # Run BASELINE for comparison
        base = run_pipeline(text, variant="BASELINE", temp_mode="fixed",
                            temperature=out.temperature_used)

        # ── Gate info ──────────────────────────────────────────
        actual_gate = "pass" if out.prompt_passed_gate else "rewrite"
        gate_correct = (actual_gate == exp_gate)

        # ── Scores ────────────────────────────────────────────
        orig_c = out.critique_original or {}
        final_c = out.critique_final or {}

        orig_scores  = orig_c.get("scores", {})
        final_scores = final_c.get("scores", {})

        orig_total  = orig_c.get("total", sum(orig_scores.values()) if orig_scores else 0)
        final_total = final_c.get("total", sum(final_scores.values()) if final_scores else 0)
        score_lift  = final_total - orig_total

        # ── LLM Judge (run TWICE for consistency check) ───────────────────
        llm_winner = llm_reason = ""
        llm_x_total = llm_y_total = 0
        llm_winner2 = ""
        judge_consistent = False
        try:
            jr1 = judge_pair(text, base.answer, out.answer)
            llm_winner  = jr1.winner
            llm_reason  = jr1.reason
            llm_x_total = total_score(jr1.X)
            llm_y_total = total_score(jr1.Y)
            # Second judge run for consistency
            jr2 = judge_pair(text, base.answer, out.answer)
            llm_winner2     = jr2.winner
            judge_consistent = (llm_winner == llm_winner2)
        except Exception as je:
            llm_winner = "error"
            llm_reason = str(je)[:80]
            judge_consistent = False

        # ── Heuristic Judge ───────────────────────────────────
        heur_winner = ""
        try:
            hj = heuristic_judge_pair(text, base.answer, out.answer)
            heur_winner = hj["winner"]
        except Exception:
            heur_winner = "error"

        judges_agree = (llm_winner == heur_winner) if (llm_winner not in ("error","") and heur_winner not in ("error","")) else False

        latency = round(time.time() - t0, 2)

        result.update({
            "route_predicted":   out.route,
            "route_correct":     out.route == route_ex,
            "actual_gate":       actual_gate,
            "gate_correct":      gate_correct,
            "enhanced_prompt":   out.enhanced_prompt,
            "rewritten":         not out.prompt_passed_gate,
            "orig_score_total":  orig_total,
            "orig_intent":       orig_scores.get("intent", ""),
            "orig_clarity":      orig_scores.get("clarity", ""),
            "orig_structure":    orig_scores.get("structure", ""),
            "orig_safety":       orig_scores.get("safety", ""),
            "final_score_total": final_total,
            "final_intent":      final_scores.get("intent", ""),
            "final_clarity":     final_scores.get("clarity", ""),
            "final_structure":   final_scores.get("structure", ""),
            "final_safety":      final_scores.get("safety", ""),
            "score_lift":        score_lift,
            "threshold_used":    orig_c.get("threshold", out.rewrite_threshold_used or ""),
            "threshold_reason":  orig_c.get("threshold_reason", ""),
            "enhance_mode":      out.enhance_mode,
            "temperature_used":  out.temperature_used,
            "baseline_response": base.answer[:300] + "..." if len(base.answer) > 300 else base.answer,
            "enhanced_response": out.answer[:300] + "..." if len(out.answer) > 300 else out.answer,
            "llm_judge_winner":  llm_winner,
            "llm_judge_reason":  llm_reason,
            "llm_score_x_total": llm_x_total,
            "llm_score_y_total": llm_y_total,
            "heur_judge_winner": heur_winner,
            "judges_agree":      judges_agree,
            "judge_consistent":  judge_consistent,
            "llm_winner_run2":   llm_winner2,
            "model_used":        getattr(__import__("gemini_client"), "LAST_MODEL_USED", "unknown"),
            "latency_seconds":   latency,
        })

    except Exception as e:
        result.update({
            "error": traceback.format_exc()[:200],
            "latency_seconds": round(time.time() - t0, 2),
        })

    return result


def write_xlsx(rows: list[dict]) -> None:
    wb = Workbook()

    # ── Sheet 1: Raw Results ──────────────────────────────────
    ws = wb.active
    ws.title = "Results"

    # Header row
    ws.append(HEADERS)
    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, col_idx)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor=COL_HEADER)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    # Data rows
    for r_idx, row in enumerate(rows, 2):
        for c_idx, h in enumerate(HEADERS, 1):
            val = row.get(h, "")
            if isinstance(val, bool):
                val = "YES" if val else "NO"
            ws.cell(r_idx, c_idx).value = val
            ws.cell(r_idx, c_idx).alignment = Alignment(wrap_text=False)

        # Color gate_correct column
        gate_col = HEADERS.index("gate_correct") + 1
        if row.get("gate_correct") is True:
            _cell_color(ws, r_idx, gate_col, COL_PASS)
        elif row.get("gate_correct") is False:
            _cell_color(ws, r_idx, gate_col, COL_FAIL)

        # Color route_correct column
        route_col = HEADERS.index("route_correct") + 1
        if row.get("route_correct") is True:
            _cell_color(ws, r_idx, route_col, COL_PASS)
        elif row.get("route_correct") is False:
            _cell_color(ws, r_idx, route_col, COL_FAIL)

        # Color score_lift
        lift_col = HEADERS.index("score_lift") + 1
        lift = row.get("score_lift", 0) or 0
        if lift > 0:
            _cell_color(ws, r_idx, lift_col, COL_PASS)
        elif lift < 0:
            _cell_color(ws, r_idx, lift_col, COL_FAIL)

        # Color judges_agree
        agree_col = HEADERS.index("judges_agree") + 1
        if row.get("judges_agree") is True:
            _cell_color(ws, r_idx, agree_col, COL_PASS)

        # Error row highlight
        if row.get("error"):
            for c in range(1, len(HEADERS) + 1):
                _cell_color(ws, r_idx, c, COL_FAIL)

    # Column widths
    col_widths = {
        "prompt_id": 8, "route_expected": 12, "route_predicted": 12,
        "route_correct": 10, "quality_tier": 10, "expected_gate": 10,
        "actual_gate": 10, "gate_correct": 10, "original_prompt": 40,
        "enhanced_prompt": 40, "rewritten": 8,
        "orig_score_total": 10, "final_score_total": 10, "score_lift": 8,
        "threshold_used": 10, "threshold_reason": 35,
        "llm_judge_winner": 12, "llm_judge_reason": 35,
        "heur_judge_winner": 12, "judges_agree": 10,
        "baseline_response": 50, "enhanced_response": 50,
        "latency_seconds": 10, "error": 30,
    }
    for h, width in col_widths.items():
        if h in HEADERS:
            ws.column_dimensions[get_column_letter(HEADERS.index(h) + 1)].width = width

    ws.freeze_panes = "A2"

    # ── Sheet 2: Summary by Route ────────────────────────────
    ws2 = wb.create_sheet("Summary by Route")
    routes = ["SOCIAL", "QA", "TASK", "TECH", "CREATIVE"]
    s_headers = ["Route", "Total", "Rewrite Rate", "Gate Accuracy",
                 "Route Accuracy", "Avg Score Lift", "LLM Win Rate (Y)",
                 "Judge Agreement", "Avg Latency (s)"]
    ws2.append(s_headers)
    for col_idx in range(1, len(s_headers) + 1):
        ws2.cell(1, col_idx).font = Font(bold=True, color="FFFFFF")
        ws2.cell(1, col_idx).fill = PatternFill("solid", fgColor=COL_HEADER)

    for route in routes:
        r_rows = [r for r in rows if r.get("route_expected") == route and not r.get("error")]
        if not r_rows:
            continue
        n = len(r_rows)
        rewrite_rate   = sum(1 for r in r_rows if r.get("rewritten")) / n
        gate_acc       = sum(1 for r in r_rows if r.get("gate_correct") is True) / n
        route_acc      = sum(1 for r in r_rows if r.get("route_correct") is True) / n
        avg_lift       = sum((r.get("score_lift") or 0) for r in r_rows) / n
        llm_win_rate   = sum(1 for r in r_rows if r.get("llm_judge_winner") == "Y") / n
        judge_agree    = sum(1 for r in r_rows if r.get("judges_agree") is True) / n
        avg_latency    = sum((r.get("latency_seconds") or 0) for r in r_rows) / n

        ws2.append([
            route, n,
            f"{rewrite_rate:.1%}", f"{gate_acc:.1%}", f"{route_acc:.1%}",
            f"{avg_lift:+.2f}", f"{llm_win_rate:.1%}",
            f"{judge_agree:.1%}", f"{avg_latency:.1f}s",
        ])

    # Overall row
    valid = [r for r in rows if not r.get("error")]
    n_all = len(valid)
    if n_all:
        ws2.append([
            "OVERALL", n_all,
            f"{sum(1 for r in valid if r.get('rewritten'))/n_all:.1%}",
            f"{sum(1 for r in valid if r.get('gate_correct') is True)/n_all:.1%}",
            f"{sum(1 for r in valid if r.get('route_correct') is True)/n_all:.1%}",
            f"{sum((r.get('score_lift') or 0) for r in valid)/n_all:+.2f}",
            f"{sum(1 for r in valid if r.get('llm_judge_winner')=='Y')/n_all:.1%}",
            f"{sum(1 for r in valid if r.get('judges_agree') is True)/n_all:.1%}",
            f"{sum((r.get('latency_seconds') or 0) for r in valid)/n_all:.1f}s",
        ])
        last_row = ws2.max_row
        for c in range(1, len(s_headers) + 1):
            ws2.cell(last_row, c).font = Font(bold=True)

    for col in range(1, len(s_headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 18
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Summary by Quality Tier ─────────────────────
    ws3 = wb.create_sheet("Summary by Quality")
    tiers = ["clean", "borderline", "messy"]
    q_headers = ["Quality Tier", "Total", "Rewrite Rate", "Gate Accuracy",
                 "Avg Score Lift", "LLM Win Rate (Y)", "Avg Latency (s)"]
    ws3.append(q_headers)
    for col_idx in range(1, len(q_headers) + 1):
        ws3.cell(1, col_idx).font = Font(bold=True, color="FFFFFF")
        ws3.cell(1, col_idx).fill = PatternFill("solid", fgColor=COL_HEADER)

    for tier in tiers:
        t_rows = [r for r in rows if r.get("quality_tier") == tier and not r.get("error")]
        if not t_rows:
            continue
        n = len(t_rows)
        ws3.append([
            tier, n,
            f"{sum(1 for r in t_rows if r.get('rewritten'))/n:.1%}",
            f"{sum(1 for r in t_rows if r.get('gate_correct') is True)/n:.1%}",
            f"{sum((r.get('score_lift') or 0) for r in t_rows)/n:+.2f}",
            f"{sum(1 for r in t_rows if r.get('llm_judge_winner')=='Y')/n:.1%}",
            f"{sum((r.get('latency_seconds') or 0) for r in t_rows)/n:.1f}s",
        ])

    for col in range(1, len(q_headers) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 18
    ws3.freeze_panes = "A2"

    wb.save(XLSX_OUT)
    print(f"\n✅ Results saved → {XLSX_OUT}")


def write_summary_txt(rows: list[dict]) -> None:
    valid  = [r for r in rows if not r.get("error")]
    errors = [r for r in rows if r.get("error")]
    n      = len(valid)

    lines = [
        "=" * 60,
        "PEISR EVALUATION SUMMARY",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "=" * 60,
        f"Total prompts run : {len(rows)}",
        f"Successful        : {n}",
        f"Errors            : {len(errors)}",
        "",
        "── OVERALL METRICS ──────────────────────────────────",
        f"Gate accuracy     : {sum(1 for r in valid if r.get('gate_correct') is True)/n:.1%}",
        f"Route accuracy    : {sum(1 for r in valid if r.get('route_correct') is True)/n:.1%}",
        f"Rewrite rate      : {sum(1 for r in valid if r.get('rewritten'))/n:.1%}",
        f"Avg score lift    : {sum((r.get('score_lift') or 0) for r in valid)/n:+.2f} / 20",
        f"LLM win rate (Y)  : {sum(1 for r in valid if r.get('llm_judge_winner')=='Y')/n:.1%}",
        f"Judge agreement   : {sum(1 for r in valid if r.get('judges_agree') is True)/n:.1%}",
        f"Avg latency       : {sum((r.get('latency_seconds') or 0) for r in valid)/n:.1f}s",
        "",
        "── BY ROUTE ─────────────────────────────────────────",
    ]

    for route in ["SOCIAL", "QA", "TASK", "TECH", "CREATIVE"]:
        r_rows = [r for r in valid if r.get("route_expected") == route]
        if not r_rows:
            continue
        nr = len(r_rows)
        lines.append(
            f"{route:10s} n={nr:2d}  gate={sum(1 for r in r_rows if r.get('gate_correct') is True)/nr:.0%}"
            f"  rewrite={sum(1 for r in r_rows if r.get('rewritten'))/nr:.0%}"
            f"  lift={sum((r.get('score_lift') or 0) for r in r_rows)/nr:+.1f}"
            f"  Y_wins={sum(1 for r in r_rows if r.get('llm_judge_winner')=='Y')/nr:.0%}"
        )

    lines += [
        "",
        "── BY QUALITY TIER ──────────────────────────────────",
    ]
    for tier in ["clean", "borderline", "messy"]:
        t_rows = [r for r in valid if r.get("quality_tier") == tier]
        if not t_rows:
            continue
        nt = len(t_rows)
        lines.append(
            f"{tier:12s} n={nt:2d}  gate={sum(1 for r in t_rows if r.get('gate_correct') is True)/nt:.0%}"
            f"  rewrite={sum(1 for r in t_rows if r.get('rewritten'))/nt:.0%}"
            f"  lift={sum((r.get('score_lift') or 0) for r in t_rows)/nt:+.1f}"
        )

    if errors:
        lines += ["", "── ERRORS ───────────────────────────────────────────"]
        for r in errors:
            lines.append(f"  {r['prompt_id']}: {r.get('error','')[:80]}")

    lines += ["", "=" * 60]
    txt = "\n".join(lines)

    with open(TXT_OUT, "w", encoding="utf-8") as f:
        f.write(txt)

    print(txt)
    print(f"\n✅ Summary saved → {TXT_OUT}")


def main():
    parser = argparse.ArgumentParser(description="PEISR Evaluation Runner")
    parser.add_argument("--limit",   type=int,  default=None, help="Max prompts to run")
    parser.add_argument("--route",   type=str,  default=None, help="Filter by route e.g. TECH")
    parser.add_argument("--quality", type=str,  default=None, help="Filter by quality: clean|borderline|messy")
    parser.add_argument("--delay",   type=float, default=2.0,  help="Seconds between prompts (avoid rate limit)")
    args = parser.parse_args()

    prompts = EVAL_PROMPTS
    if args.route:
        prompts = [p for p in prompts if p["route"] == args.route.upper()]
    if args.quality:
        prompts = [p for p in prompts if p["quality"] == args.quality.lower()]
    if args.limit:
        prompts = prompts[:args.limit]

    print(f"\n🚀 Running PEISR eval on {len(prompts)} prompts...")
    print(f"   Delay between prompts: {args.delay}s")
    print(f"   Output: {XLSX_OUT}\n")

    results = []
    for i, p in enumerate(prompts, 1):
        print(f"[{i:02d}/{len(prompts)}] {p['id']} ({p['route']}/{p['quality']}) — {p['text'][:60]}...")
        row = run_single(p)

        status = "✅" if not row.get("error") else "❌"
        gate_ok = "✓gate" if row.get("gate_correct") else "✗gate"
        route_ok = "✓route" if row.get("route_correct") else "✗route"
        rewritten = "REWRITE" if row.get("rewritten") else "DIRECT"
        lift = row.get("score_lift", 0) or 0
        winner = row.get("llm_judge_winner", "?")

        print(f"       {status} {gate_ok} {route_ok} | {rewritten} | lift={lift:+d} | winner={winner} | {row.get('latency_seconds',0):.1f}s")

        results.append(row)

        # Delay to avoid rate limiting
        if i < len(prompts):
            time.sleep(args.delay)

    print(f"\n📊 Writing results...")
    write_xlsx(results)
    write_summary_txt(results)


if __name__ == "__main__":
    main()
