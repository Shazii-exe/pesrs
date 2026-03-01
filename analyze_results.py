# analyze_results.py
# ─────────────────────────────────────────────────────────────────────
# Pulls all data from Supabase (or SQLite fallback), computes every
# metric needed for the paper, and writes a paper-ready Excel workbook
# with 6 sheets + embedded charts.
#
# Usage:
#   python analyze_results.py                  # full export
#   python analyze_results.py --min-ratings 1  # include unrated rows too
#
# Output:
#   peisr_results_<timestamp>.xlsx
# ─────────────────────────────────────────────────────────────────────

from __future__ import annotations

import argparse
import json
import os
import sqlite3
import sys
from datetime import datetime
from typing import Any

import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import (Alignment, Border, Font, GradientFill,
                              PatternFill, Side)
from openpyxl.utils import get_column_letter

load_dotenv()

# ── colours ──────────────────────────────────────────────────────────
NAVY   = "1E3A5F"
BLUE   = "2563EB"
CYAN   = "0EA5E9"
GREEN  = "10B981"
AMBER  = "F59E0B"
RED    = "EF4444"
LIGHT  = "EFF6FF"
WHITE  = "FFFFFF"
MUTED  = "94A3B8"
DARK   = "0F172A"

DB_PATH = "peisr_runs.db"

# ═════════════════════════════════════════════════════════════════════
# DATA FETCHING
# ═════════════════════════════════════════════════════════════════════

def fetch_all_data() -> tuple[pd.DataFrame, pd.DataFrame]:
    """Return (comparisons_df, inline_ratings_df)."""

    # Try Supabase first
    try:
        from supabase import create_client
        url = os.getenv("SUPABASE_URL", "")
        key = os.getenv("SUPABASE_KEY", "")
        if url and key:
            sb = create_client(url, key)
            print("✅ Connected to Supabase")

            comp_res = sb.table("comparisons").select("*").execute()
            rate_res = sb.table("inline_ratings").select("*").execute()

            comp_df = pd.DataFrame(comp_res.data or [])
            rate_df = pd.DataFrame(rate_res.data or [])
            print(f"   comparisons : {len(comp_df)} rows")
            print(f"   inline_ratings: {len(rate_df)} rows")
            return comp_df, rate_df
    except Exception as e:
        print(f"⚠️  Supabase failed ({e}), falling back to SQLite…")

    # SQLite fallback
    if not os.path.exists(DB_PATH):
        print(f"❌  No SQLite DB found at {DB_PATH}")
        sys.exit(1)

    conn = sqlite3.connect(DB_PATH)
    comp_df = pd.read_sql("SELECT * FROM comparisons", conn)
    try:
        rate_df = pd.read_sql("SELECT * FROM inline_ratings", conn)
    except Exception:
        rate_df = pd.DataFrame()
    conn.close()
    print(f"✅ Loaded from SQLite: {len(comp_df)} comparisons, {len(rate_df)} inline ratings")
    return comp_df, rate_df


# ═════════════════════════════════════════════════════════════════════
# DATA PREPARATION
# ═════════════════════════════════════════════════════════════════════

def prepare(comp_df: pd.DataFrame, rate_df: pd.DataFrame) -> pd.DataFrame:
    if comp_df.empty:
        print("❌  No comparison data found. Make sure users have rated responses.")
        sys.exit(1)

    df = comp_df.copy()

    # Parse JSON columns
    for col in ["original_prompt_critique_json", "enhanced_prompt_critique_json",
                "response_llm_judge_json", "response_heuristic_judge_json",
                "original_prompt_heuristic_json", "enhanced_prompt_heuristic_json"]:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: _safe_json(x))

    # Scores
    df["orig_score"]  = df["original_prompt_critique_json"].apply(lambda x: _total(x))
    df["final_score"] = df["enhanced_prompt_critique_json"].apply(lambda x: _total(x))
    df["score_lift"]  = df["final_score"] - df["orig_score"]

    # Gate correctness: rewritten prompts should have had low scores; direct ones high
    # We define correct gate as: rewritten when orig_score < threshold, passed when >=
    df["threshold"]   = df["original_prompt_critique_json"].apply(lambda x: x.get("threshold", 15) if isinstance(x, dict) else 15)
    df["gate_correct"] = df.apply(
        lambda r: (r["rewritten"] and r["orig_score"] < r["threshold"]) or
                  (not r["rewritten"] and r["orig_score"] >= r["threshold"]),
        axis=1,
    )

    # LLM judge winner
    df["llm_winner"] = df["response_llm_judge_json"].apply(
        lambda x: x.get("winner", "") if isinstance(x, dict) else ""
    )
    df["heur_winner"] = df["response_heuristic_judge_json"].apply(
        lambda x: x.get("winner", "") if isinstance(x, dict) else ""
    )
    df["judges_agree"] = df["llm_winner"] == df["heur_winner"]

    # Enhanced wins
    df["enhanced_won"] = df["llm_winner"].isin(["Y", "ENHANCED"])

    # Merge inline ratings
    if not rate_df.empty and "comparison_id" in rate_df.columns:
        rate_sub = rate_df[["comparison_id", "stars", "pick", "notes"]].rename(
            columns={"stars": "inline_stars", "pick": "inline_pick", "notes": "inline_notes"}
        )
        df = df.merge(rate_sub, on="comparison_id", how="left")
    else:
        df["inline_stars"] = None
        df["inline_pick"]  = None
        df["inline_notes"] = None

    # Normalise route column
    route_col = "route_predicted" if "route_predicted" in df.columns else "route"
    df["route"] = df[route_col].str.upper().str.strip()

    # Timestamp
    if "ts" in df.columns:
        df["ts"] = pd.to_datetime(df["ts"], errors="coerce")

    return df


def _safe_json(x: Any) -> dict:
    if isinstance(x, dict):
        return x
    if isinstance(x, str):
        try:
            return json.loads(x)
        except Exception:
            return {}
    return {}


def _total(x: Any) -> float:
    if isinstance(x, dict):
        scores = x.get("scores", x)
        if isinstance(scores, dict):
            return float(sum(v for v in scores.values() if isinstance(v, (int, float))))
        total = x.get("total", x.get("score", 0))
        return float(total)
    return 0.0


# ═════════════════════════════════════════════════════════════════════
# METRIC COMPUTATIONS
# ═════════════════════════════════════════════════════════════════════

ROUTES = ["SOCIAL", "QA", "TASK", "TECH", "CREATIVE"]

def compute_overall(df: pd.DataFrame) -> dict:
    total       = len(df)
    rewritten   = df["rewritten"].sum()
    direct      = total - rewritten
    gate_acc    = df["gate_correct"].mean() * 100
    avg_lift    = df.loc[df["rewritten"], "score_lift"].mean() if rewritten else 0
    win_rate    = df.loc[df["rewritten"], "enhanced_won"].mean() * 100 if rewritten else 0
    judge_agree = df["judges_agree"].mean() * 100
    rated       = df["inline_stars"].notna().sum()
    avg_stars   = df["inline_stars"].mean() if rated else 0

    return {
        "Total Prompts":          total,
        "Rewritten":              int(rewritten),
        "Direct (Gate Pass)":     int(direct),
        "Rewrite Rate (%)":       round(rewritten / total * 100, 1),
        "Gate Accuracy (%)":      round(gate_acc, 1),
        "Avg Score Lift":         round(avg_lift, 2),
        "Enhanced Win Rate (%)":  round(win_rate, 1),
        "Judge Agreement (%)":    round(judge_agree, 1),
        "Rated by Humans":        int(rated),
        "Avg Human Stars":        round(avg_stars, 2),
    }


def compute_by_route(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for route in ROUTES:
        sub = df[df["route"] == route]
        if sub.empty:
            rows.append({"Route": route, "n": 0})
            continue
        n         = len(sub)
        rw        = sub["rewritten"].sum()
        rw_sub    = sub[sub["rewritten"]]
        rows.append({
            "Route":               route,
            "n":                   n,
            "Rewrite Rate (%)":    round(rw / n * 100, 1),
            "Gate Accuracy (%)":   round(sub["gate_correct"].mean() * 100, 1),
            "Avg Orig Score":      round(sub["orig_score"].mean(), 1),
            "Avg Final Score":     round(sub["final_score"].mean(), 1),
            "Avg Score Lift":      round(rw_sub["score_lift"].mean(), 2) if len(rw_sub) else 0,
            "Enhanced Win Rate (%)": round(rw_sub["enhanced_won"].mean() * 100, 1) if len(rw_sub) else 0,
            "Judge Agreement (%)": round(sub["judges_agree"].mean() * 100, 1),
            "Avg Human Stars":     round(sub["inline_stars"].mean(), 2) if sub["inline_stars"].notna().any() else "—",
        })
    return pd.DataFrame(rows)


def compute_human_ratings(df: pd.DataFrame) -> pd.DataFrame:
    rated = df[df["inline_stars"].notna()].copy()
    if rated.empty:
        return pd.DataFrame(columns=["Route", "n Rated", "Avg Stars", "Enhanced Preferred (%)", "Original Preferred (%)"])

    rows = []
    for route in ROUTES + ["ALL"]:
        sub = rated if route == "ALL" else rated[rated["route"] == route]
        if sub.empty:
            continue
        n          = len(sub)
        avg_stars  = sub["inline_stars"].mean()
        enh_pref   = (sub["inline_pick"].str.upper().isin(["ENHANCED", "Y"])).sum() / n * 100
        orig_pref  = (sub["inline_pick"].str.upper().isin(["ORIGINAL", "X"])).sum() / n * 100
        tie_pref   = (sub["inline_pick"].str.upper() == "TIE").sum() / n * 100
        rows.append({
            "Route":                    route,
            "n Rated":                  n,
            "Avg Stars":                round(avg_stars, 2),
            "Enhanced Preferred (%)":   round(enh_pref, 1),
            "Original Preferred (%)":   round(orig_pref, 1),
            "Tie (%)":                  round(tie_pref, 1),
        })
    return pd.DataFrame(rows)


def compute_score_distribution(df: pd.DataFrame) -> pd.DataFrame:
    bins   = list(range(0, 22, 2))
    labels = [f"{b}–{b+2}" for b in bins[:-1]]
    orig_hist  = pd.cut(df["orig_score"],  bins=bins, labels=labels, right=False).value_counts().sort_index()
    final_hist = pd.cut(df["final_score"], bins=bins, labels=labels, right=False).value_counts().sort_index()
    return pd.DataFrame({"Score Range": labels, "Original Prompts": orig_hist.values, "After Enhancement": final_hist.values})


def compute_raw(df: pd.DataFrame) -> pd.DataFrame:
    cols = [
        "ts", "human_rater", "route", "user_input",
        "rewritten", "orig_score", "final_score", "score_lift",
        "threshold", "gate_correct",
        "llm_winner", "heur_winner", "judges_agree", "enhanced_won",
        "inline_stars", "inline_pick", "inline_notes",
        "model_used", "original_prompt", "enhanced_prompt",
    ]
    available = [c for c in cols if c in df.columns]
    return df[available].copy()


# ═════════════════════════════════════════════════════════════════════
# EXCEL WRITING
# ═════════════════════════════════════════════════════════════════════

def write_excel(overall: dict, by_route: pd.DataFrame, human: pd.DataFrame,
                score_dist: pd.DataFrame, raw: pd.DataFrame, df: pd.DataFrame,
                path: str):

    with pd.ExcelWriter(path, engine="openpyxl") as writer:

        # ── Sheet 1: Summary ──────────────────────────────────
        summ_df = pd.DataFrame([
            {"Metric": k, "Value": v} for k, v in overall.items()
        ])
        summ_df.to_excel(writer, sheet_name="Summary", index=False, startrow=2)

        # ── Sheet 2: By Route ─────────────────────────────────
        by_route.to_excel(writer, sheet_name="By Route", index=False, startrow=2)

        # ── Sheet 3: Human Ratings ────────────────────────────
        human.to_excel(writer, sheet_name="Human Ratings", index=False, startrow=2)

        # ── Sheet 4: Score Distribution ───────────────────────
        score_dist.to_excel(writer, sheet_name="Score Distribution", index=False, startrow=2)

        # ── Sheet 5: Session Timeline ─────────────────────────
        if "ts" in df.columns and df["ts"].notna().any():
            timeline = (df.groupby(df["ts"].dt.date)
                         .agg(prompts=("comparison_id", "count"),
                              rewrites=("rewritten", "sum"),
                              avg_stars=("inline_stars", "mean"))
                         .reset_index()
                         .rename(columns={"ts": "date"}))
            timeline.to_excel(writer, sheet_name="Timeline", index=False, startrow=2)
        else:
            pd.DataFrame({"note": ["No timestamp data yet"]}).to_excel(
                writer, sheet_name="Timeline", index=False, startrow=2)

        # ── Sheet 6: Raw Data ─────────────────────────────────
        raw.to_excel(writer, sheet_name="Raw Data", index=False, startrow=2)

    # ── Post-process: styling + charts ───────────────────────
    wb = load_workbook(path)
    _style_summary(wb["Summary"], overall)
    _style_table(wb["By Route"],        by_route,     "By Route")
    _style_table(wb["Human Ratings"],   human,        "Human Ratings")
    _style_table(wb["Score Distribution"], score_dist, "Score Distribution")
    if "Timeline" in wb.sheetnames:
        _style_table(wb["Timeline"], None, "Timeline")
    _style_raw(wb["Raw Data"])
    _add_charts(wb, by_route, score_dist)
    wb.save(path)
    print(f"\n✅ Excel saved → {path}")


# ── Styling helpers ───────────────────────────────────────────────

def _hdr_fill(color): return PatternFill("solid", fgColor=color)
def _font(bold=False, color="000000", size=11):
    return Font(name="Calibri", bold=bold, color=color, size=size)
def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def _center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_summary(ws, overall: dict):
    ws.insert_rows(1, 2)
    ws["A1"] = "PEISR — Evaluation Summary"
    ws["A1"].font   = Font(name="Calibri", bold=True, size=14, color=WHITE)
    ws["A1"].fill   = _hdr_fill(NAVY)
    ws["A1"].alignment = _center()
    ws.merge_cells("A1:B1")

    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = _font(color=MUTED, size=10)

    for row in ws.iter_rows(min_row=3, max_row=3):
        for cell in row:
            cell.font  = _font(bold=True, color=WHITE)
            cell.fill  = _hdr_fill(NAVY)
            cell.alignment = _center()
            cell.border = _border()

    highlight_metrics = {"Gate Accuracy (%)": GREEN, "Enhanced Win Rate (%)": BLUE,
                         "Avg Score Lift": AMBER, "Avg Human Stars": CYAN}
    for row in ws.iter_rows(min_row=4):
        label = str(row[0].value or "")
        for cell in row:
            cell.border = _border()
            cell.alignment = _center()
        if label in highlight_metrics:
            for cell in row:
                cell.fill = _hdr_fill(LIGHT)
                cell.font = _font(bold=True)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20


def _style_table(ws, df, title: str):
    ws.insert_rows(1, 2)
    ws["A1"] = f"PEISR — {title}"
    ws["A1"].font      = Font(name="Calibri", bold=True, size=13, color=WHITE)
    ws["A1"].fill      = _hdr_fill(NAVY)
    ws["A1"].alignment = _center()

    # Merge A1 across all used columns
    if df is not None and not df.empty:
        last_col = get_column_letter(len(df.columns))
        try:
            ws.merge_cells(f"A1:{last_col}1")
        except Exception:
            pass

    for row in ws.iter_rows(min_row=3, max_row=3):
        for cell in row:
            cell.font  = _font(bold=True, color=WHITE)
            cell.fill  = _hdr_fill(BLUE)
            cell.alignment = _center()
            cell.border = _border()

    for row in ws.iter_rows(min_row=4):
        for i, cell in enumerate(row):
            cell.border    = _border()
            cell.alignment = _center()
            if i == 0:
                cell.font = _font(bold=True)
            if row[0].row % 2 == 0:
                cell.fill = _hdr_fill(LIGHT)

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)


def _style_raw(ws):
    ws.insert_rows(1, 2)
    ws["A1"] = "PEISR — Raw Data Export"
    ws["A1"].font  = Font(name="Calibri", bold=True, size=13, color=WHITE)
    ws["A1"].fill  = _hdr_fill(DARK)
    ws["A1"].alignment = _center()

    for row in ws.iter_rows(min_row=3, max_row=3):
        for cell in row:
            cell.font  = _font(bold=True, color=WHITE)
            cell.fill  = _hdr_fill(DARK)
            cell.alignment = _center()
            cell.border = _border()

    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.border    = _border()
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if row[0].row % 2 == 0:
                cell.fill = _hdr_fill("F8FAFC")

    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 22

    ws.freeze_panes = "A4"


def _add_charts(wb, by_route: pd.DataFrame, score_dist: pd.DataFrame):
    """Add charts into a dedicated Charts sheet."""
    cs = wb.create_sheet("Charts")
    cs["A1"] = "PEISR — Charts (auto-generated)"
    cs["A1"].font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    cs["A1"].fill = PatternFill("solid", fgColor=NAVY)

    route_ws = wb["By Route"]

    # ── Chart 1: Rewrite Rate by Route (bar) ──────────────────
    if not by_route.empty and "Rewrite Rate (%)" in by_route.columns:
        try:
            col_idx = by_route.columns.tolist().index("Rewrite Rate (%)") + 1
            chart1  = BarChart()
            chart1.type    = "col"
            chart1.title   = "Rewrite Rate by Route (%)"
            chart1.y_axis.title = "Rewrite Rate (%)"
            chart1.x_axis.title = "Route"
            chart1.style   = 10
            chart1.width   = 18
            chart1.height  = 12

            data = Reference(route_ws, min_col=col_idx+1, max_col=col_idx+1,
                             min_row=3, max_row=3+len(by_route))
            cats = Reference(route_ws, min_col=1, min_row=4, max_row=3+len(by_route))
            chart1.add_data(data, titles_from_data=True)
            chart1.set_categories(cats)
            chart1.series[0].graphicalProperties.solidFill = BLUE
            cs.add_chart(chart1, "A3")
        except Exception as e:
            print(f"  ⚠️  Chart 1 skipped: {e}")

    # ── Chart 2: Gate Accuracy by Route (bar) ─────────────────
    if not by_route.empty and "Gate Accuracy (%)" in by_route.columns:
        try:
            col_idx = by_route.columns.tolist().index("Gate Accuracy (%)") + 1
            chart2  = BarChart()
            chart2.type    = "col"
            chart2.title   = "Gate Accuracy by Route (%)"
            chart2.y_axis.title = "Gate Accuracy (%)"
            chart2.style   = 10
            chart2.width   = 18
            chart2.height  = 12

            data = Reference(route_ws, min_col=col_idx+1, max_col=col_idx+1,
                             min_row=3, max_row=3+len(by_route))
            cats = Reference(route_ws, min_col=1, min_row=4, max_row=3+len(by_route))
            chart2.add_data(data, titles_from_data=True)
            chart2.set_categories(cats)
            chart2.series[0].graphicalProperties.solidFill = GREEN
            cs.add_chart(chart2, "J3")
        except Exception as e:
            print(f"  ⚠️  Chart 2 skipped: {e}")

    # ── Chart 3: Score distribution before/after (bar) ────────
    if not score_dist.empty:
        try:
            dist_ws = wb["Score Distribution"]
            chart3  = BarChart()
            chart3.type    = "col"
            chart3.title   = "Prompt Score Distribution: Before vs After Enhancement"
            chart3.y_axis.title = "Number of Prompts"
            chart3.x_axis.title = "Score Range (/20)"
            chart3.style   = 10
            chart3.width   = 22
            chart3.height  = 14
            chart3.grouping = "clustered"

            data = Reference(dist_ws, min_col=2, max_col=3,
                             min_row=3, max_row=3+len(score_dist))
            cats = Reference(dist_ws, min_col=1, min_row=4, max_row=3+len(score_dist))
            chart3.add_data(data, titles_from_data=True)
            chart3.set_categories(cats)
            if chart3.series:
                chart3.series[0].graphicalProperties.solidFill = AMBER
            if len(chart3.series) > 1:
                chart3.series[1].graphicalProperties.solidFill = GREEN
            cs.add_chart(chart3, "A23")
        except Exception as e:
            print(f"  ⚠️  Chart 3 skipped: {e}")

    # ── Chart 4: Enhanced win rate by route (bar) ─────────────
    if not by_route.empty and "Enhanced Win Rate (%)" in by_route.columns:
        try:
            col_idx = by_route.columns.tolist().index("Enhanced Win Rate (%)") + 1
            chart4  = BarChart()
            chart4.type    = "col"
            chart4.title   = "Enhanced Response Win Rate by Route (%)"
            chart4.y_axis.title = "Win Rate (%)"
            chart4.style   = 10
            chart4.width   = 18
            chart4.height  = 12

            data = Reference(route_ws, min_col=col_idx+1, max_col=col_idx+1,
                             min_row=3, max_row=3+len(by_route))
            cats = Reference(route_ws, min_col=1, min_row=4, max_row=3+len(by_route))
            chart4.add_data(data, titles_from_data=True)
            chart4.set_categories(cats)
            chart4.series[0].graphicalProperties.solidFill = CYAN
            cs.add_chart(chart4, "J23")
        except Exception as e:
            print(f"  ⚠️  Chart 4 skipped: {e}")


# ═════════════════════════════════════════════════════════════════════
# PRINT SUMMARY TO TERMINAL
# ═════════════════════════════════════════════════════════════════════

def print_summary(overall: dict, by_route: pd.DataFrame, human: pd.DataFrame):
    print("\n" + "═"*55)
    print("  PEISR EVALUATION RESULTS")
    print("═"*55)
    for k, v in overall.items():
        print(f"  {k:<30} {v}")

    print("\n  BY ROUTE:")
    print(f"  {'Route':<12} {'n':>4} {'Rewrite%':>9} {'GateAcc%':>9} {'Lift':>6} {'Win%':>6} {'Stars':>6}")
    print("  " + "-"*55)
    for _, row in by_route.iterrows():
        print(f"  {row.get('Route',''):<12} "
              f"{row.get('n',0):>4} "
              f"{row.get('Rewrite Rate (%)','—'):>9} "
              f"{row.get('Gate Accuracy (%)','—'):>9} "
              f"{row.get('Avg Score Lift','—'):>6} "
              f"{row.get('Enhanced Win Rate (%)','—'):>6} "
              f"{row.get('Avg Human Stars','—'):>6}")

    if not human.empty:
        print("\n  HUMAN RATINGS:")
        for _, row in human.iterrows():
            print(f"  {row.get('Route',''):<12} "
                  f"n={row.get('n Rated',0)}  "
                  f"stars={row.get('Avg Stars','—')}  "
                  f"enh_pref={row.get('Enhanced Preferred (%)','—')}%")
    print("═"*55 + "\n")


# ═════════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="PEISR results analyzer")
    parser.add_argument("--min-ratings", type=int, default=0,
                        help="Minimum inline_stars value to include (0 = include unrated)")
    args = parser.parse_args()

    print("\n🔍 PEISR Results Analyzer")
    print("─"*40)

    comp_df, rate_df = fetch_all_data()
    df               = prepare(comp_df, rate_df)

    if args.min_ratings > 0:
        df = df[df["inline_stars"] >= args.min_ratings]
        print(f"   Filtered to rows with stars >= {args.min_ratings}: {len(df)} rows")

    overall    = compute_overall(df)
    by_route   = compute_by_route(df)
    human      = compute_human_ratings(df)
    score_dist = compute_score_distribution(df)
    raw        = compute_raw(df)

    print_summary(overall, by_route, human)

    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = f"peisr_results_{ts}.xlsx"
    write_excel(overall, by_route, human, score_dist, raw, df, path)
    print(f"📊 Open {path} to see all tables and charts.")
    print("   Copy numbers from Summary + By Route sheets into your paper.\n")


if __name__ == "__main__":
    main()
