from __future__ import annotations

import json
from datetime import datetime
from typing import Any, Dict, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

DEFAULT_XLSX = "peisr_runs.xlsx"

# ---------- Sheets & headers ----------
RUNS_SHEET = "runs"
RUN_HEADERS = [
    "timestamp",
    "run_id",
    "variant",
    "route_predicted",
    "enhance_mode",
    "temperature_used",
    "rewrite_threshold_used",
    "original_prompt",
    "enhanced_prompt",
    "response_text",
    # Judge fields (optional)
    "judge_model",
    "judge_score_overall",
    "judge_intent",
    "judge_clarity",
    "judge_structure",
    "judge_safety",
    "judge_notes",
    # Human fields (optional)
    "human_rater",
    "human_score_overall",
    "human_pick",
    "human_notes",
]

COMPARE_SHEET = "comparisons"
COMPARE_HEADERS = [
    "timestamp",
    "comparison_id",
    "human_rater",
    "user_input",
    "route_predicted",
    "temperature_used",
    "rewrite_threshold_used",
    "rewritten",
    # Left (original)
    "original_prompt",
    "original_response",
    "original_prompt_critique_json",
    # Right (rewritten / enhanced)
    "enhanced_prompt",
    "enhanced_response",
    "enhanced_prompt_critique_json",
    # Human ratings
    "human_score_original",
    "human_score_enhanced",
    "human_pick",  # ORIGINAL | ENHANCED | TIE
    "human_notes",
]


# ---------- Workbook helpers ----------
def _ensure_sheet(wb, title: str, headers: list[str]) -> Worksheet:
    if title in wb.sheetnames:
        ws = wb[title]
        # If empty, write headers
        if ws.max_row == 0 or ws.cell(1, 1).value is None:
            ws.append(headers)
        else:
            # If headers missing/mismatched, we do not rewrite (avoid breaking old data)
            existing = [c.value for c in ws[1]]
            if existing != headers:
                # Add any missing columns at end
                existing_set = set(existing)
                for h in headers:
                    if h not in existing_set:
                        ws.cell(1, ws.max_column + 1).value = h
        return ws

    ws = wb.create_sheet(title)
    ws.append(headers)
    return ws


def _load_or_create(path: str):
    try:
        wb = load_workbook(path)
    except FileNotFoundError:
        wb = Workbook()
        # Remove default sheet to avoid confusion
        if wb.active and wb.active.title == "Sheet":
            wb.remove(wb.active)
    return wb


def _acquire_lock(lock_path: str) -> Optional[int]:
    """
    Naive cross-platform lock using atomic file creation.
    Returns fd if lock acquired, else None.
    """
    import os
    try:
        fd = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_RDWR)
        return fd
    except FileExistsError:
        return None


def _release_lock(fd: int, lock_path: str):
    import os
    try:
        os.close(fd)
    finally:
        try:
            os.remove(lock_path)
        except FileNotFoundError:
            pass


# ---------- Public API for experiments ----------
def append_run(
    *,
    run_id: str,
    variant: str,
    route_predicted: str,
    enhance_mode: str,
    temperature_used: float,
    rewrite_threshold_used: Optional[int],
    original_prompt: str,
    enhanced_prompt: str,
    response_text: str,
    judge: Optional[Dict[str, Any]] = None,
    human: Optional[Dict[str, Any]] = None,
    path: str = DEFAULT_XLSX,
) -> None:
    """Append one pipeline run to the RUNS sheet."""
    judge = judge or {}
    human = human or {}

    lock_path = path + ".lock"
    fd = _acquire_lock(lock_path)
    if fd is None:
        # If locked, fall back to retry-free safe behavior: just raise.
        raise RuntimeError("Excel file is currently locked by another user/process. Try again.")

    try:
        wb = _load_or_create(path)
        ws = _ensure_sheet(wb, RUNS_SHEET, RUN_HEADERS)

        ws.append([
            datetime.now().isoformat(timespec="seconds"),
            run_id,
            variant,
            route_predicted,
            enhance_mode,
            float(temperature_used),
            rewrite_threshold_used,
            original_prompt,
            enhanced_prompt,
            response_text,
            judge.get("judge_model"),
            judge.get("overall"),
            judge.get("intent"),
            judge.get("clarity"),
            judge.get("structure"),
            judge.get("safety"),
            judge.get("notes"),
            human.get("rater"),
            human.get("overall"),
            human.get("pick"),
            human.get("notes"),
        ])

        wb.save(path)
    finally:
        _release_lock(fd, lock_path)


def set_human_rating(
    *,
    run_id: str,
    variant: str,
    human_rater: str,
    human_score_overall: int,
    human_pick: str,
    human_notes: str = "",
    path: str = DEFAULT_XLSX,
) -> bool:
    """Update human rating for an existing run row. Returns True if updated."""
    lock_path = path + ".lock"
    fd = _acquire_lock(lock_path)
    if fd is None:
        raise RuntimeError("Excel file is currently locked by another user/process. Try again.")

    try:
        wb = _load_or_create(path)
        ws = _ensure_sheet(wb, RUNS_SHEET, RUN_HEADERS)

        headers = [c.value for c in ws[1]]
        col = {h: i + 1 for i, h in enumerate(headers)}

        for r in range(2, ws.max_row + 1):
            if ws.cell(r, col["run_id"]).value == run_id and ws.cell(r, col["variant"]).value == variant:
                ws.cell(r, col["human_rater"]).value = human_rater
                ws.cell(r, col["human_score_overall"]).value = int(human_score_overall)
                ws.cell(r, col["human_pick"]).value = human_pick
                ws.cell(r, col["human_notes"]).value = human_notes
                wb.save(path)
                return True

        return False
    finally:
        _release_lock(fd, lock_path)


# ---------- Public API for Streamlit comparisons ----------
def append_comparison(
    *,
    comparison_id: str,
    human_rater: str,
    user_input: str,
    route_predicted: str,
    temperature_used: float,
    rewrite_threshold_used: Optional[int],
    rewritten: bool,
    original_prompt: str,
    original_response: str,
    original_prompt_critique: Dict[str, Any],
    enhanced_prompt: str,
    enhanced_response: str,
    enhanced_prompt_critique: Dict[str, Any],
    human_score_original: int,
    human_score_enhanced: int,
    human_pick: str,
    human_notes: str = "",
    path: str = DEFAULT_XLSX,
) -> None:
    """Append a single UI comparison row to the COMPARISONS sheet."""
    lock_path = path + ".lock"
    fd = _acquire_lock(lock_path)
    if fd is None:
        raise RuntimeError("Excel file is currently locked by another user/process. Try again.")

    try:
        wb = _load_or_create(path)
        ws = _ensure_sheet(wb, COMPARE_SHEET, COMPARE_HEADERS)

        ws.append([
            datetime.now().isoformat(timespec="seconds"),
            comparison_id,
            human_rater,
            user_input,
            route_predicted,
            float(temperature_used),
            rewrite_threshold_used,
            bool(rewritten),
            original_prompt,
            original_response,
            json.dumps(original_prompt_critique, ensure_ascii=False),
            enhanced_prompt,
            enhanced_response,
            json.dumps(enhanced_prompt_critique, ensure_ascii=False),
            int(human_score_original),
            int(human_score_enhanced),
            human_pick,
            human_notes,
        ])
        wb.save(path)
    finally:
        _release_lock(fd, lock_path)


def workbook_bytes(path: str = DEFAULT_XLSX) -> Optional[bytes]:
    """Return workbook bytes for Streamlit download button."""
    import os
    if not os.path.exists(path):
        return None
    with open(path, "rb") as f:
        return f.read()
